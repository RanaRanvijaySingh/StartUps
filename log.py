import os
from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime

def isNotBlank (myString):
    return bool(myString and myString.strip())

hrs = input(">>>> Log hour: ") 

lastCommitCommand = 'git log -1 --pretty=%B'
branchCommand = 'git name-rev --name-only HEAD'

message = os.popen(lastCommitCommand).read().strip()
branch = os.popen(branchCommand).read().strip()

date = datetime.now().strftime('%d-%m-%Y')
month = datetime.now().strftime('%B')
year = datetime.now().strftime('%Y')
day = datetime.now().strftime('%d')

logsDir = '../../logs/'
logFileName = month + "_" + year + ".xlsx"
logSheet = logsDir + logFileName

if not os.path.exists(logsDir):
	os.makedirs(logsDir)

if os.path.exists(logSheet):
	print(">>>> Loading existing file: " + logSheet)
	wb = load_workbook(filename = logSheet)
else:
	print(">>>> Creating new file: " + logSheet)
	wb = Workbook() 

ws = wb.active

loggedMessage = ws['C' + day].value
loggedHour = ws['D' + day].value
logHour = 0
logMessage = ""

if loggedHour is None:
	logHour = int(hrs)
else:
	logHour = int(hrs) + int(loggedHour)

if loggedMessage is None:
	logMessage = message
else:
	logMessage = message + "\n" + loggedMessage


ws['A' + day] = date
ws['B' + day] = branch
ws['C' + day] = logMessage
ws['D' + day] = logHour
wb.save(logSheet)
print(">>>> Logging: " + date + " " + branch + " " + message + " " + hrs)
print(">>>> DONE")

